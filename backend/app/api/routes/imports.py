import csv
import io
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from app.api.deps import require_roles
from app.db.session import get_db
from app.models.entities import User
from app.schemas.leads import LeadCreate, ContactIn
from app.services.lead_upsert import upsert_lead

router = APIRouter(prefix="/imports", tags=["imports"])
MAX_BYTES = 5 * 1024 * 1024
MAX_ROWS = 20_000

ALIASES = {
    "business_name": ["business_name", "name", "نام", "نام کسب و کار", "نام کسب‌وکار"],
    "industry": ["industry", "صنف", "حوزه"], "province": ["province", "استان"],
    "city": ["city", "شهر"], "address": ["address", "آدرس"], "website": ["website", "وبسایت", "وب‌سایت"],
    "phone": ["phone", "mobile", "تلفن", "موبایل"], "whatsapp": ["whatsapp", "واتساپ"],
    "instagram": ["instagram", "اینستاگرام"], "telegram": ["telegram", "تلگرام"],
    "eitaa": ["eitaa", "ایتا"], "rubika": ["rubika", "روبیکا"],
}


def _canonical(row: dict) -> dict:
    normalized = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
    out = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            if alias.lower() in normalized and normalized[alias.lower()] not in (None, ""):
                out[field] = str(normalized[alias.lower()]).strip(); break
    return out


def _to_lead(row: dict, source: str) -> LeadCreate | None:
    r = _canonical(row)
    if not r.get("business_name"):
        return None
    contacts=[]
    for field, channel in [("phone","PHONE"),("whatsapp","WHATSAPP"),("instagram","INSTAGRAM_HANDLE"),
                           ("telegram","TELEGRAM_HANDLE"),("eitaa","EITAA_HANDLE"),("rubika","RUBIKA_HANDLE")]:
        if r.get(field): contacts.append(ContactIn(channel=channel, value=r[field]))
    if r.get("website"): contacts.append(ContactIn(channel="WEB", value=r["website"]))
    return LeadCreate(business_name=r["business_name"], industry=r.get("industry"), province=r.get("province"),
                      city=r.get("city"), address=r.get("address"), website=r.get("website"), source=source,
                      contacts=contacts)


def _read_csv(data: bytes):
    text = data.decode("utf-8-sig")
    yield from csv.DictReader(io.StringIO(text))


def _read_xlsx(data: bytes):
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try: headers = [str(x).strip() if x is not None else "" for x in next(rows)]
    except StopIteration: return
    for values in rows:
        yield dict(zip(headers, values))


@router.post("/leads")
async def import_leads(file: UploadFile = File(...), db: AsyncSession = Depends(get_db), user: User = Depends(require_roles("admin", "marketing", "supervisor"))):
    data = await file.read(MAX_BYTES + 1)
    if len(data) > MAX_BYTES: raise HTTPException(413, "file exceeds 5 MB")
    name=(file.filename or "").lower()
    if name.endswith(".csv"):
        rows=_read_csv(data)
    elif name.endswith(".xlsx"):
        rows=_read_xlsx(data)
    else:
        raise HTTPException(415, "only CSV and XLSX are supported")
    processed=created_or_merged=skipped=0
    for row in rows:
        processed += 1
        if processed > MAX_ROWS: raise HTTPException(413, f"row limit {MAX_ROWS} exceeded")
        payload=_to_lead(row, source="IMPORT")
        if not payload: skipped += 1; continue
        await upsert_lead(db, payload, actor_user_id=user.id)
        created_or_merged += 1
    await db.commit()
    return {"processed":processed,"created_or_merged":created_or_merged,"skipped":skipped}
